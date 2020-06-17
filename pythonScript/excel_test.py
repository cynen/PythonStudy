import openpyxl


# 创建excel文件.
def create_excel():
    wb=openpyxl.Workbook()
    wb.create_sheet("test_case")
    wb.save("D:\\case.xlsx") # 指定路径.

# 打开Excel文件,并读取数据.
def open_excel():
    wb=openpyxl.load_workbook("cities.xlsx") # 打开指定的exel
    sh=wb['tb_cities'] # 选择sheet
    # 读取数据
    # 按行读取数据. 去掉第一行表头.
    # print(list(sh.rows)[1:])
    print(len(list(sh.rows))) # 获取总条数.
    for row in list(sh.rows)[1:]:
        id = row[0].value
        cityid = row[1].value
        city = row[2].value
        print(id,cityid,city)

if __name__ == "__main__":
    create_excel()
    # open_excel()