import openpyxl
from openpyxl import Workbook

# 打开Excel文件,并写数据.
def read_excel():
    wb=openpyxl.load_workbook("D:\\case.xlsx") # 打开指定的exel
    sheet=wb['test_case']
    data = sheet.cell(1,1).value # 获取单元格值.
    print(type(data))

# 打开Excel文件,并写数据.
def write_to_excel():
    wb=openpyxl.load_workbook("D:\\case.xlsx") # 打开指定的exel
    sheet=wb['test_case']
    sheet.cell(1,1).value='欣怡2'
    sheet.append([2,3,4,5,6])
    sheet.append([2,3,4,5,6])
    sheet.append([2,3,4,5,6])
    print("写入...")
    wb.save("D:\\case.xlsx") # 保存的和打开的excel不能是同一个.
    # wb.save("D:\\case.xlsx") # 保存的和打开的excel不能是同一个.


if __name__ == "__main__":
    write_to_excel()
